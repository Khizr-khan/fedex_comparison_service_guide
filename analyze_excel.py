"""
STEP 0 — Run this FIRST before writing ingest_excel.py.
Prints sheet names and first 10 rows of both Excel files so you can
confirm zone column positions and service block start rows.

Usage:
    python analyze_excel.py \
        --excel2026 "Service Guide/FedEx_Standard_List_Rates_2026.xlsx" \
        --excel2025 "Service Guide/FedEx_Standard_List_Rates_2025.xlsx"
"""

import argparse
import openpyxl

def analyze(path: str, label: str):
    print(f"\n{'='*60}")
    print(f"FILE: {label}  →  {path}")
    print(f"{'='*60}")
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  ── Sheet: '{sheet_name}'  (max_row={ws.max_row}, max_col={ws.max_column})")
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i > 15:
                break
            print(f"    Row {i:>3}: {list(row)}")
        print()

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel2026", required=True)
    parser.add_argument("--excel2025", required=True)
    args = parser.parse_args()

    analyze(args.excel2026, "2026")
    analyze(args.excel2025, "2025")

if __name__ == "__main__":
    main()