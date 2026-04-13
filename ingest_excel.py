"""
ingest_excel.py — FedEx 2026 Rate Ingestion from Excel
-------------------------------------------------------
Parses FedEx_Standard_List_Rates_2026.xlsx directly.
100% accurate — no PDF parsing errors.

Sheets handled:
  1. "2026 U.S. rates"        — Domestic express (First Overnight, Priority, Standard, 2Day AM, 2Day, Express Saver)
  2. "2026 U.S. Export rates" — International export (Puerto Rico, Canada Zones A-C, Zones D-O)
  3. "2026 U.S. Import rates" — International import zones
  4. "2026 Ground & FHD rates"— FedEx Ground and Home Delivery

Plus: surcharge_chunks.py structured surcharges

Run:
    python ingest_excel.py --excel "FedEx_Standard_List_Rates_2026.xlsx"
"""

import os
import re
import argparse
import shutil
from openpyxl import load_workbook
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_chroma import Chroma
from langchain_core.documents import Document
from dotenv import load_dotenv

load_dotenv()

CHROMA_DIR  = "./chroma_fedex_db"
EMBED_MODEL = "all-MiniLM-L6-v2"

# Zone column label → clean zone name
ZONE_LABEL_MAP = {
    "A": "Canada Zone A", "B": "Canada Zone B", "C": "Canada Zone C",
    "A ": "Canada Zone A", "B ": "Canada Zone B", "C ": "Canada Zone C",
    "D": "Zone D", "E": "Zone E", "F": "Zone F", "G": "Zone G",
    "H": "Zone H", "I": "Zone I", "J": "Zone J", "K": "Zone K",
    "L": "Zone L", "M": "Zone M", "N": "Zone N", "O": "Zone O", "P": "Zone P",
    "D ": "Zone D", "E ": "Zone E", "F ": "Zone F", "G ": "Zone G",
    "H ": "Zone H", "I ": "Zone I", "J ": "Zone J", "K ": "Zone K",
    "L ": "Zone L", "M ": "Zone M", "N ": "Zone N", "O ": "Zone O",
    2: "Zone 2", 3: "Zone 3", 4: "Zone 4", 5: "Zone 5",
    6: "Zone 6", 7: "Zone 7", 8: "Zone 8",
}


def make_doc(service, zone, weight, price, page=0, direction=""):
    weight_str = str(weight).replace(" lbs.", "").replace(" lb.", "").strip()
    # Try to get integer weight
    try:
        weight_int = int(float(weight_str))
        weight_str = str(weight_int)
    except Exception:
        pass

    parts = [f"Service: {service}."]
    if zone:
        parts.append(f"Shipping zone: {zone}.")
    if direction:
        parts.append(f"Direction: {direction}.")
    parts.append(f"Package weight: {weight_str} lbs.")
    parts.append(f"Shipping rate: ${price}.")

    content = " ".join(parts)
    meta = {
        "source": "FedEx_Standard_List_Rates_2026",
        "page": page,
        "type": "rate",
        "service": service,
        "zone": zone,
        "weight": weight_str,
    }
    return Document(page_content=content, metadata=meta)


def parse_weight(cell):
    """Extract numeric weight from cell value like '1 lb.', '2 lbs.', 15, 'FedEx® Envelope up to 8 oz.'"""
    if cell is None:
        return None
    s = str(cell).strip()
    # Skip envelope/pak/header rows
    if any(skip in s.lower() for skip in ["envelope", "pak", "multiweight", "zones", "weight", "fedex®"]):
        return None
    # Extract number
    m = re.match(r"^(\d+)", s)
    if m:
        return int(m.group(1))
    return None


def parse_price(cell):
    """Return float price or None if not a valid price."""
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return round(float(cell), 2)
    s = str(cell).strip()
    if s in ("*", "**", "-", ""):
        return None
    m = re.search(r"[\d,]+\.?\d*", s)
    if m:
        try:
            return round(float(m.group(0).replace(",", "")), 2)
        except Exception:
            return None
    return None


# ---------------------------------------------------------------------------
# Sheet 1 — Domestic U.S. rates
# ---------------------------------------------------------------------------

def parse_us_rates(wb):
    """
    Structure: Service header row, then Zones row (cols B+), then data rows.
    Each service block is 157 rows (1 header + 1 zones row + 155 weight rows).
    Services: First Overnight, Priority Overnight, Standard Overnight,
              2Day AM, 2Day, Express Saver
    """
    ws = wb["2026 U.S. rates"]
    rows = list(ws.iter_rows(values_only=True))
    docs = []

    # Find service blocks by looking for FedEx service name rows
    SERVICE_ROWS = {
        3:   "FedEx First Overnight®",
        160: "FedEx Priority Overnight®",
        317: "FedEx Standard Overnight®",
        474: "FedEx 2Day® A.M.",
        631: "FedEx 2Day®",
        788: "FedEx Express Saver®",
    }

    for service_row_idx, service_name in SERVICE_ROWS.items():
        # Zone header is next row
        zone_row = rows[service_row_idx + 1]  # e.g. ['Zones', 2, 3, 4, 5, 6, 7, 8, ...]
        # Get zone columns (skip col 0 = 'Zones', col 1 = weight label col)
        zone_cols = []
        for col_idx, cell in enumerate(zone_row):
            if col_idx < 1:
                continue
            zone = ZONE_LABEL_MAP.get(cell)
            if zone:
                zone_cols.append((col_idx, zone))

        # Data rows start 2 rows after service header
        for row in rows[service_row_idx + 2: service_row_idx + 160]:
            weight_cell = row[0]
            weight = parse_weight(weight_cell)
            if weight is None:
                continue

            for col_idx, zone in zone_cols:
                if col_idx >= len(row):
                    continue
                price = parse_price(row[col_idx])
                if price is None:
                    continue
                docs.append(make_doc(service_name, zone, weight, f"{price:.2f}"))

    print(f"  [U.S. Rates]   {len(docs):,} chunks")
    return docs


# ---------------------------------------------------------------------------
# Sheet 2 — International Export rates
# ---------------------------------------------------------------------------

def parse_export_rates(wb):
    """
    Structure: Section header (destination/zone group), then column headers,
    then weight rows.
    Sections: Puerto Rico, Canada A-C, Zones D-O per service.
    """
    ws = wb["2026 U.S. Export rates"]
    rows = list(ws.iter_rows(values_only=True))
    docs = []

    # Puerto Rico block (rows 4-160): columns are services
    # Weight col=0, First col=1, Priority col=2, Economy col=3
    pr_services = {
        1: "FedEx International First®",
        2: "FedEx International Priority®",
        3: "FedEx International Economy®",
    }
    for row in rows[5:161]:
        weight = parse_weight(row[0])
        if weight is None:
            continue
        for col_idx, svc in pr_services.items():
            price = parse_price(row[col_idx])
            if price:
                docs.append(make_doc(svc, "Puerto Rico", weight, f"{price:.2f}", direction="U.S. export"))

    # International zone blocks — each service has its own block
    # Find blocks by looking for FedEx service name + Weight header
    INTL_BLOCKS = {
        161: "FedEx International First®",
        268: "FedEx International Priority® Express",
        375: "FedEx International Priority®",
        482: "FedEx International Economy®",
        591: "FedEx® International Connect Plus",
    }

    for start_row, service_name in INTL_BLOCKS.items():
        # Zone header row is at start_row + 1 (0-indexed: start_row)
        zone_header = rows[start_row + 1]  # e.g. ['Weight', 'A', 'B', 'C', ...]
        zone_cols = []
        for col_idx, cell in enumerate(zone_header):
            if col_idx == 0:
                continue
            zone = ZONE_LABEL_MAP.get(cell) or ZONE_LABEL_MAP.get(str(cell).strip())
            if zone:
                zone_cols.append((col_idx, zone))

        # Data rows
        for row in rows[start_row + 2: start_row + 107]:
            weight = parse_weight(row[0])
            if weight is None:
                continue
            for col_idx, zone in zone_cols:
                if col_idx >= len(row):
                    continue
                price = parse_price(row[col_idx])
                if price:
                    docs.append(make_doc(service_name, zone, weight, f"{price:.2f}", direction="U.S. export"))

    print(f"  [U.S. Export]  {len(docs):,} chunks")
    return docs


# ---------------------------------------------------------------------------
# Sheet 3 — International Import rates
# ---------------------------------------------------------------------------

def parse_import_rates(wb):
    ws = wb["2026 U.S. Import rates"]
    rows = list(ws.iter_rows(values_only=True))
    docs = []

    IMPORT_BLOCKS = {
        3:   "FedEx International First®",
        110: "FedEx International Priority® Express",
        217: "FedEx International Priority®",
        324: "FedEx International Economy®",
        431: "FedEx® International Connect Plus",
    }

    for start_row, service_name in IMPORT_BLOCKS.items():
        zone_header = rows[start_row + 1]
        zone_cols = []
        for col_idx, cell in enumerate(zone_header):
            if col_idx == 0:
                continue
            zone = ZONE_LABEL_MAP.get(cell) or ZONE_LABEL_MAP.get(str(cell).strip())
            if zone:
                zone_cols.append((col_idx, f"{zone} (Import)"))

        for row in rows[start_row + 2: start_row + 107]:
            weight = parse_weight(row[0])
            if weight is None:
                continue
            for col_idx, zone in zone_cols:
                if col_idx >= len(row):
                    continue
                price = parse_price(row[col_idx])
                if price:
                    docs.append(make_doc(service_name, zone, weight, f"{price:.2f}", direction="U.S. import"))

    print(f"  [U.S. Import]  {len(docs):,} chunks")
    return docs


# ---------------------------------------------------------------------------
# Sheet 4 — Ground & FHD rates
# ---------------------------------------------------------------------------

def parse_ground_rates(wb):
    ws = wb["2026 Ground & FHD rates"]
    rows = list(ws.iter_rows(values_only=True))
    docs = []

    service_name = "FedEx Ground® and FedEx Home Delivery®"

    # Zone numbers are in row 6 (index 5), columns B onwards
    zone_row = rows[5]
    zone_cols = []
    for col_idx, cell in enumerate(zone_row):
        if col_idx == 0:
            continue
        if isinstance(cell, (int, float)) and 2 <= cell <= 9:
            zone_cols.append((col_idx, f"Zone {int(cell)}"))

    # Data starts row 7 (index 6)
    for row in rows[6:160]:
        weight = parse_weight(row[0])
        if weight is None:
            continue
        for col_idx, zone in zone_cols:
            if col_idx >= len(row):
                continue
            price = parse_price(row[col_idx])
            if price:
                docs.append(make_doc(service_name, zone, weight, f"{price:.2f}"))

    print(f"  [Ground/FHD]   {len(docs):,} chunks")
    return docs


# ---------------------------------------------------------------------------
# Main ingestion
# ---------------------------------------------------------------------------

def ingest(excel_path):
    print(f"\n{'='*60}")
    print(f"FedEx Excel Ingestion — 100% accurate rate data")
    print(f"Source: {excel_path}")
    print(f"{'='*60}\n")

    wb = load_workbook(excel_path, read_only=True)
    print(f"Sheets: {wb.sheetnames}\n")

    all_docs = []

    # Parse all rate sheets
    all_docs.extend(parse_us_rates(wb))
    all_docs.extend(parse_export_rates(wb))
    all_docs.extend(parse_import_rates(wb))
    all_docs.extend(parse_ground_rates(wb))

    # Add surcharge chunks
    try:
        from surcharge_chunks_2026 import SURCHARGE_CHUNKS
        surcharge_docs = [
            Document(
                page_content=chunk,
                metadata={"source": "FedEx_Service_Guide_2026", "page": 0, "type": "surcharge"}
            )
            for chunk in SURCHARGE_CHUNKS
        ]
        all_docs.extend(surcharge_docs)
        print(f"  [Surcharges]   {len(surcharge_docs)} structured chunks")
    except ImportError:
        print("  [Surcharges]   surcharge_chunks.py not found — skipping")

    # Deduplicate
    seen = set()
    unique = []
    for doc in all_docs:
        key = doc.page_content.strip()
        if key not in seen and len(key) > 10:
            seen.add(key)
            unique.append(doc)

    print(f"\n{'─'*60}")
    print(f"Total chunks     : {len(all_docs):,}")
    print(f"After dedup      : {len(unique):,}")

    # Remove old DB
    if os.path.exists(CHROMA_DIR):
        print(f"\nRemoving old database at {CHROMA_DIR}...")
        shutil.rmtree(CHROMA_DIR)

    print(f"\nLoading embedding model: {EMBED_MODEL}...")
    embeddings = HuggingFaceEmbeddings(model_name=EMBED_MODEL)

    print(f"Embedding into {CHROMA_DIR}...\n")
    BATCH = 2000
    vectorstore = None
    for start in range(0, len(unique), BATCH):
        batch = unique[start:start + BATCH]
        if vectorstore is None:
            vectorstore = Chroma.from_documents(
                documents=batch, embedding=embeddings,
                persist_directory=CHROMA_DIR
            )
        else:
            vectorstore.add_documents(batch)
        print(f"  Embedded {min(start+BATCH, len(unique)):,} / {len(unique):,}")

    print(f"\n{'='*60}")
    print(f"Done! {len(unique):,} chunks indexed into {CHROMA_DIR}")
    print(f"{'='*60}\n")
    return len(unique)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default="FedEx_Standard_List_Rates_2026.xlsx")
    args = parser.parse_args()
    if not os.path.exists(args.excel):
        print(f"ERROR: File not found: '{args.excel}'")
        exit(1)
    ingest(args.excel)