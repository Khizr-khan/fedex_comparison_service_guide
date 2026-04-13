"""
diagnose3.py — run: python diagnose3.py
Checks the export sheet structure deeply and Connect Plus detection.
"""
import openpyxl, re

EXCEL = "Service Guide/FedEx_Standard_List_Rates_2026.xlsx"
wb = openpyxl.load_workbook(EXCEL, data_only=True)

# ── 1. Export sheet — print ALL non-empty col-A rows ─────────────────────────
print("=" * 60)
print("EXPORT SHEET — every non-None col A (all rows)")
print("=" * 60)
ws = wb['2026 U.S. Export rates']
total_rows = ws.max_row
print(f"Total rows: {total_rows}")
section_rows = []
for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    col_a = row[0]
    if col_a is not None and not isinstance(col_a, (int, float)):
        # String rows — these are section headers or weight labels
        val = str(col_a).strip()
        # Skip pure weight strings like '1 lb.'
        if not re.match(r'^\d+\s*lbs?\.?$', val, re.IGNORECASE):
            section_rows.append((i, val, list(row[1:5])))

for rownum, val, rest in section_rows:
    print(f"  Row {rownum:>4}: {repr(val[:80])}  | B-E: {rest}")

# ── 2. Connect Plus — search for it in ALL sheets ────────────────────────────
print("\n" + "=" * 60)
print("CONNECT PLUS — searching all sheets for service name")
print("=" * 60)
for sheet_name in wb.sheetnames:
    ws2 = wb[sheet_name]
    for i, row in enumerate(ws2.iter_rows(values_only=True), start=1):
        for cell in row:
            if cell and "connect" in str(cell).lower():
                print(f"  Sheet '{sheet_name}' Row {i}: {repr(str(cell)[:100])}")

# ── 3. Ground sheet — check for duplicate zone labels ────────────────────────
print("\n" + "=" * 60)
print("GROUND SHEET — all zone header rows (rows 5-6)")
print("=" * 60)
ws3 = wb['2026 Ground & FHD rates']
rows = list(ws3.iter_rows(values_only=True))
print(f"Row 5: {rows[4]}")
print(f"Row 6: {rows[5]}")
# Show full zone labels from row 5 and numbers from row 6
print("\nZone label mapping (row5 label → row6 number):")
for col_i, (label, num) in enumerate(zip(rows[4][1:], rows[5][1:]), start=1):
    if label is not None or num is not None:
        print(f"  col {col_i+1}: label={repr(label)} → zone_num={num}")